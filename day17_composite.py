from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from day16_screeners import PresetScreeners


OUTPUT_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "year",
    "roe",
    "roce",
    "npm",
    "fcf_cagr_5yr",
    "cfo_pat_ratio",
    "fcf_positive_flag",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "de_ratio",
    "icr",
    "profitability_score",
    "cash_quality_score",
    "growth_score",
    "leverage_score",
    "composite_quality_score",
    "preset",
]

METRIC_COLUMNS = {
    "roe",
    "roce",
    "npm",
    "fcf_cagr_5yr",
    "cfo_pat_ratio",
    "fcf_positive_flag",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "de_ratio",
    "icr",
}

PRESET_LABELS = {
    "quality_compounder": "Quality Compounder",
    "value_pick": "Value Pick",
    "growth_accelerator": "Growth Accelerator",
    "dividend_champion": "Dividend Champion",
    "debt_free_blue_chip": "Debt-Free Blue Chip",
    "turnaround_watch": "Turnaround Watch",
}


class CompositeScoreExporter:
    """Score the latest company records and export preset workbooks."""

    def __init__(self, data: pd.DataFrame | str | Path):
        if isinstance(data, pd.DataFrame):
            frame = data.copy()
        else:
            frame = pd.read_csv(Path(data))

        frame.columns = [
            column.strip().lower().replace(" ", "_") for column in frame.columns
        ]
        self.screeners = PresetScreeners(frame)
        self.data = self.screeners.latest.copy()
        self._require_columns(
            {
                "company_id",
                "company_name",
                "broad_sector",
                "roe",
                "roce",
                "npm",
                "fcf_cagr_5yr",
                "cfo_pat_ratio",
                "free_cash_flow_crore",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "de_ratio",
                "icr",
            }
        )
        self.scored = self._score_by_sector(self.data)

    def _require_columns(self, columns: Iterable[str]) -> None:
        missing = sorted(set(columns).difference(self.data.columns))
        if missing:
            raise ValueError("Missing required columns: " + ", ".join(missing))

    @staticmethod
    def _scale_p10_p90(values: pd.Series, higher_is_better: bool = True) -> pd.Series:
        numeric = pd.to_numeric(values, errors="coerce")
        low = numeric.quantile(0.10)
        high = numeric.quantile(0.90)
        if pd.isna(low) or pd.isna(high) or high == low:
            score = pd.Series(50.0, index=values.index)
        else:
            capped = numeric.clip(lower=low, upper=high)
            score = (capped - low).div(high - low).mul(100)
        if not higher_is_better:
            score = 100 - score
        return score.fillna(50.0).clip(0, 100)

    @classmethod
    def _score_by_sector(cls, frame: pd.DataFrame) -> pd.DataFrame:
        result = frame.copy()
        result["fcf_positive_flag"] = result["free_cash_flow_crore"].gt(0).astype(int)
        score_specs = {
            "roe": True,
            "roce": True,
            "npm": True,
            "fcf_cagr_5yr": True,
            "cfo_pat_ratio": True,
            "fcf_positive_flag": True,
            "revenue_cagr_5yr": True,
            "pat_cagr_5yr": True,
            "de_ratio": False,
            "icr": True,
        }
        for metric, higher_is_better in score_specs.items():
            result[f"_{metric}_score"] = result.groupby("broad_sector", group_keys=False)[
                metric
            ].transform(
                lambda values: cls._scale_p10_p90(values, higher_is_better)
            )

        result["profitability_score"] = (
            result["_roe_score"] * 0.15
            + result["_roce_score"] * 0.10
            + result["_npm_score"] * 0.10
        )
        result["cash_quality_score"] = (
            result["_fcf_cagr_5yr_score"] * 0.15
            + result["_cfo_pat_ratio_score"] * 0.10
            + result["_fcf_positive_flag_score"] * 0.05
        )
        result["growth_score"] = (
            result["_revenue_cagr_5yr_score"] * 0.10
            + result["_pat_cagr_5yr_score"] * 0.10
        )
        result["leverage_score"] = (
            result["_de_ratio_score"] * 0.10
            + result["_icr_score"] * 0.05
        )
        result["composite_quality_score"] = result[
            ["profitability_score", "cash_quality_score", "growth_score", "leverage_score"]
        ].sum(axis=1).clip(0, 100)
        return result

    def _preset_result(self, preset_key: str) -> pd.DataFrame:
        preset_frame = getattr(self.screeners, preset_key)()
        result = preset_frame.merge(
            self.scored[OUTPUT_COLUMNS[:-1]],
            on=["company_id", "company_name"],
            how="left",
            suffixes=("_preset", ""),
        )
        result["preset"] = PRESET_LABELS[preset_key]
        return result[OUTPUT_COLUMNS].sort_values(
            "composite_quality_score", ascending=False, kind="stable"
        ).reset_index(drop=True)

    def results(self) -> dict[str, pd.DataFrame]:
        return {key: self._preset_result(key) for key in PRESET_LABELS}

    @staticmethod
    def _threshold_status(row: pd.Series, preset_key: str, column: str) -> bool:
        rules = {
            "quality_compounder": {
                "roe": row["roe"] > 15,
                "de_ratio": row["de_ratio"] < 1,
                "fcf_positive_flag": row["fcf_positive_flag"] == 1,
                "revenue_cagr_5yr": row["revenue_cagr_5yr"] > 10,
            },
            "value_pick": {
                "de_ratio": row["de_ratio"] < 2,
            },
            "growth_accelerator": {
                "pat_cagr_5yr": row["pat_cagr_5yr"] > 20,
                "revenue_cagr_5yr": row["revenue_cagr_5yr"] > 15,
                "de_ratio": row["de_ratio"] < 2,
            },
            "dividend_champion": {
                "fcf_positive_flag": row["fcf_positive_flag"] == 1,
            },
            "debt_free_blue_chip": {
                "de_ratio": row["de_ratio"] == 0,
                "roe": row["roe"] > 12,
            },
            "turnaround_watch": {
                "revenue_cagr_5yr": row["revenue_cagr_5yr"] > 10,
                "fcf_positive_flag": row["fcf_positive_flag"] == 1,
            },
        }
        return rules[preset_key].get(column, True)

    def export(self, output_path: str | Path = "output/screener_output.xlsx") -> Path:
        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        results = self.results()
        with pd.ExcelWriter(destination, engine="openpyxl") as writer:
            for preset_key, result in results.items():
                result.to_excel(
                    writer,
                    sheet_name=PRESET_LABELS[preset_key][:31],
                    index=False,
                )

        workbook = load_workbook(destination)
        green = PatternFill(fill_type="solid", fgColor="C6EFCE")
        red = PatternFill(fill_type="solid", fgColor="FFC7CE")
        for preset_key, result in results.items():
            sheet = workbook[PRESET_LABELS[preset_key][:31]]
            columns = {cell.value: cell.column for cell in sheet[1]}
            for row_number, (_, row) in enumerate(result.iterrows(), start=2):
                for column in METRIC_COLUMNS:
                    cell = sheet.cell(row=row_number, column=columns[column])
                    cell.fill = (
                        green
                        if self._threshold_status(row, preset_key, column)
                        else red
                    )
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        workbook.save(destination)
        return destination
