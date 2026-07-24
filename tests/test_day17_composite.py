from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from day16_screeners import PresetScreeners
from day17_composite import OUTPUT_COLUMNS, CompositeScoreExporter
from tests.test_day16_screeners import make_universe


def make_scored_universe() -> pd.DataFrame:
    frame = make_universe()
    frame["broad_sector"] = frame["company_id"].mod(2).map({0: "Industrials", 1: "Technology"})
    frame["roce"] = frame["roe"] + 5
    frame["npm"] = frame["roe"] / 2
    frame["fcf_cagr_5yr"] = frame["revenue_cagr_5yr"]
    frame["cfo_pat_ratio"] = 1.0 + frame["roe"] / 100
    frame["icr"] = 5.0 - frame["de_ratio"]
    return frame


def test_composite_score_is_sector_relative_and_bounded():
    exporter = CompositeScoreExporter(make_scored_universe())
    scored = exporter.scored

    assert scored["composite_quality_score"].between(0, 100).all()
    assert scored["profitability_score"].between(0, 35).all()
    assert scored["cash_quality_score"].between(0, 30).all()
    assert scored["growth_score"].between(0, 20).all()
    assert scored["leverage_score"].between(0, 15).all()

    sector_max = scored.groupby("broad_sector")["roe"].transform("max")
    max_rows = scored.loc[scored["roe"].eq(sector_max)]
    assert (max_rows["_roe_score"] == 100).all()


def test_export_has_six_sheets_twenty_columns_sorted_and_colored(tmp_path: Path):
    exporter = CompositeScoreExporter(make_scored_universe())
    destination = exporter.export(tmp_path / "screener_output.xlsx")
    workbook = load_workbook(destination)

    assert len(workbook.sheetnames) == 6
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        assert sheet.max_column == len(OUTPUT_COLUMNS) == 20
        scores = [cell.value for cell in sheet["S"]][1:]
        assert scores == sorted(scores, reverse=True)
        fills = {
            sheet.cell(row=row, column=column).fill.fgColor.rgb
            for row in range(2, sheet.max_row + 1)
            for column in range(1, sheet.max_column + 1)
        }
        assert any(color.endswith("C6EFCE") for color in fills)

    failing_row = exporter.scored.iloc[0].copy()
    failing_row["roe"] = 0
    assert not exporter._threshold_status(
        failing_row, "quality_compounder", "roe"
    )


def test_day16_latest_rows_remain_one_per_company():
    assert len(PresetScreeners(make_scored_universe()).latest) == 92
